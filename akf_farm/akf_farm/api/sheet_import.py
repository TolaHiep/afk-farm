import frappe
from akf_farm.engine.frequency import parse_frequency

SCOPE_MAP = {"dùng chung": "shared", "toàn bộ lô": "shared", "theo cây": "per_crop"}


def import_rows(process_name, crop, rows):
    """Tạo Cultivation Process + steps từ list dict (cột Bước, Mô tả, Công/ha, Tần suất, Phạm vi)."""
    steps = []
    for i, r in enumerate(rows, 1):
        ftype, fval = parse_frequency(str(r.get("Tần suất", "")))
        scope = SCOPE_MAP.get(str(r.get("Phạm vi", "")).strip().lower(), "per_crop")
        steps.append({
            "step": r.get("Bước", i), "description": r["Mô tả"],
            "mandays_per_ha": r.get("Công/ha", 0), "frequency_type": ftype,
            "frequency_value": fval, "scope": scope,
        })
    doc = frappe.get_doc({
        "doctype": "Cultivation Process", "process_name": process_name, "crop": crop, "steps": steps,
    }).insert()
    return doc.name


@frappe.whitelist()
def import_file(process_name, crop, file_url):
    import openpyxl

    path = frappe.get_site_path(file_url.lstrip("/"))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, [c.value for c in row], strict=False)) for row in ws.iter_rows(min_row=2)]
    return import_rows(process_name, crop, rows)
