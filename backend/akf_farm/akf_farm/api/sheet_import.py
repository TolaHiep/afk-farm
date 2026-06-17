import frappe
from akf_farm.engine.frequency import parse_frequency

SCOPE_MAP = {"dùng chung": "shared", "toàn bộ lô": "shared", "theo cây": "per_crop"}


def _truthy(v):
    return str(v or "").strip().lower() in ("x", "có", "co", "1", "yes", "true", "y")


def _offset(v):
    if v is None or str(v).strip() == "":
        return -1
    try:
        return int(float(str(v).strip()))
    except ValueError:
        return -1


def import_rows(process_name, crop, rows):
    """Tạo Cultivation Process + steps từ list dict (cột Bước, Mô tả, Công/ha, Tần suất, Phạm vi)."""
    steps = []
    for i, r in enumerate(rows, 1):
        ftype, fval, ftimes = parse_frequency(str(r.get("Tần suất", "")))
        scope = SCOPE_MAP.get(str(r.get("Phạm vi", "")).strip().lower(), "per_crop")
        steps.append({
            "step": r.get("Bước", i), "description": r["Mô tả"],
            "mandays_per_ha": r.get("Công/ha", 0), "frequency_type": ftype,
            "frequency_value": fval, "times_per_period": ftimes, "scope": scope,
            "require_photo": 1 if _truthy(r.get("Yêu cầu ảnh", "")) else 0,
            "offset_days": _offset(r.get("Bắt đầu sau (ngày)", "")),
        })
    doc = frappe.get_doc({
        "doctype": "Cultivation Process", "process_name": process_name, "crop": crop, "steps": steps,
    }).insert()
    return doc.name


def _num(s):
    s = str(s).replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return 0


def parse_markdown_tables(text):
    """Tách các bảng quy trình trong docs/quy-trinh-canh-tac.md theo từng cây.

    Trả dict {crop: [rows]} với rows = list dict (Bước, Mô tả, Công/ha, Tần suất, Phạm vi).
    Bỏ qua dòng marker lặp ("Quay về Bước 1") và dòng không có lịch (Phạm vi rỗng/—).
    """
    out = {}
    crop = None
    for line in text.splitlines():
        ls = line.strip()
        if ls.startswith("## ") and "Sâm" in ls:
            crop, out[crop] = "Sâm", []
            continue
        if ls.startswith("## ") and "Gấc" in ls:
            crop, out[crop] = "Gấc", []
            continue
        if not crop or not ls.startswith("|"):
            continue
        cells = [c.strip() for c in ls.strip("|").split("|")]
        if len(cells) < 5:
            continue
        if cells[0].lower() in ("bước", "") or set(cells[0]) <= {"-", " "}:
            continue  # dòng header hoặc dòng kẻ "---"
        buoc, mota, congha, tansuat, phamvi = cells[:5]
        if "quay về" in mota.lower():
            continue  # marker lặp chu kỳ
        if phamvi in ("—", "-", ""):
            continue  # không có lịch
        out[crop].append({
            "Bước": int(buoc) if buoc.isdigit() else buoc, "Mô tả": mota,
            "Công/ha": _num(congha), "Tần suất": tansuat, "Phạm vi": phamvi,
        })
    return out


def import_from_markdown(path=None, replace=True):
    """Đọc file quy trình (.md) và tạo Cultivation Process cho Sâm + Gấc. Idempotent (replace)."""
    import os

    if not path:
        # docs/ nằm ở gốc repo; app được mount/symlink, đi lên từ app_path
        app_path = frappe.get_app_path("akf_farm")  # .../akf_farm/akf_farm
        path = os.path.abspath(os.path.join(app_path, "..", "..", "docs", "quy-trinh-canh-tac.md"))
    with open(path, encoding="utf-8") as f:
        tables = parse_markdown_tables(f.read())
    created = {}
    for crop, rows in tables.items():
        name = f"Quy trình {crop}"
        if frappe.db.exists("Cultivation Process", name):
            if not replace:
                continue
            frappe.delete_doc("Cultivation Process", name, force=True)
        import_rows(name, crop, rows)
        created[name] = len(rows)
    frappe.db.commit()
    return created


def parse_workbook(wb):
    """Đọc workbook mẫu: B1=tên quy trình, B2=cây, dòng header 'Bước', rồi các bước.

    Trả (name, crop, rows) với rows = list dict key tiếng Việt khớp import_rows.
    """
    ws = wb.active
    name = str(ws["B1"].value or "").strip()
    crop = str(ws["B2"].value or "").strip()

    header_row, headers = None, []
    for row in ws.iter_rows(min_row=1):
        if str(row[0].value or "").strip().lower() == "bước":
            header_row = row[0].row
            headers = [str(c.value or "").strip() for c in row]
            break
    if header_row is None:
        frappe.throw("Không tìm thấy dòng tiêu đề 'Bước' trong file.")

    rows = []
    for vals in ws.iter_rows(min_row=header_row + 1, values_only=True):
        rec = {headers[i]: vals[i] for i in range(len(headers)) if i < len(vals)}
        if not str(rec.get("Mô tả", "") or "").strip():
            continue
        rows.append(rec)

    if not name:
        frappe.throw("Thiếu Tên quy trình (ô B1).")
    if crop not in ("Gấc", "Sâm"):
        frappe.throw("Cây phải là 'Gấc' hoặc 'Sâm' (ô B2).")
    if not rows:
        frappe.throw("File chưa có bước nào (cột Mô tả trống).")
    return name, crop, rows


@frappe.whitelist()
def import_process_excel(file_b64, replace=0):
    """Nhận file xlsx (base64), đọc + tạo Cultivation Process. Trùng tên: cần replace=1 để ghi đè."""
    import base64
    import io
    import openpyxl

    try:
        raw = base64.b64decode(file_b64)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        frappe.throw("File không đọc được — hãy dùng đúng mẫu .xlsx.")

    name, crop, rows = parse_workbook(wb)
    if frappe.db.exists("Cultivation Process", name):
        if not int(replace or 0):
            return {"exists": True, "name": name}
        frappe.delete_doc("Cultivation Process", name, force=True)
    import_rows(name, crop, rows)
    return {"exists": False, "name": name, "crop": crop, "steps": len(rows)}


@frappe.whitelist()
def process_template():
    """Sinh file Excel mẫu để người dùng tải về, điền rồi upload lại."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quy trình"
    ws["A1"], ws["B1"] = "Tên quy trình", "Quy trình Gấc"
    ws["A2"], ws["B2"] = "Cây", "Gấc"
    header = ["Bước", "Mô tả", "Công/ha", "Tần suất", "Phạm vi", "Yêu cầu ảnh", "Bắt đầu sau (ngày)"]
    for col, h in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=h)
    samples = [
        [1, "Đào hố trồng 60x60cm", 2, "1 lần/20 năm", "Theo cây", "", ""],
        [2, "Bón phân nước định kỳ", 2, "60 ngày/lần", "Theo cây", "", ""],
        [3, "Tưới mát", "", "2 lần/ngày", "Dùng chung", "", ""],
        [4, "Kiểm tra sâu, bệnh", "", "Hàng ngày", "Dùng chung", "x", ""],
    ]
    for ri, r in enumerate(samples, start=5):
        for ci, v in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=v)

    bio = io.BytesIO()
    wb.save(bio)
    frappe.response["filename"] = "mau-quy-trinh.xlsx"
    frappe.response["filecontent"] = bio.getvalue()
    frappe.response["type"] = "binary"


@frappe.whitelist()
def import_file(process_name, crop, file_url):
    import openpyxl

    path = frappe.get_site_path(file_url.lstrip("/"))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, [c.value for c in row], strict=False)) for row in ws.iter_rows(min_row=2)]
    return import_rows(process_name, crop, rows)
