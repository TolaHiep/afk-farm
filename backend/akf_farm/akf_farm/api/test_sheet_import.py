import frappe
from frappe.tests.utils import FrappeTestCase
from akf_farm.api.sheet_import import import_rows
import openpyxl
from akf_farm.api import sheet_import


def _build_wb(name="Quy trình Test", crop="Gấc", data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "Tên quy trình", name
    ws["A2"], ws["B2"] = "Cây", crop
    header = ["Bước", "Mô tả", "Công/ha", "Tần suất", "Phạm vi", "Yêu cầu ảnh"]
    for col, h in enumerate(header, start=1):
        ws.cell(row=4, column=col, value=h)
    if data is None:
        data = [
            [1, "Đào hố", 2, "1 lần/20 năm", "Theo cây", ""],
            [2, "Tưới mát", "", "2 lần/ngày", "Dùng chung", "x"],
        ]
    for ri, r in enumerate(data, start=5):
        for ci, v in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=v)
    return wb


class TestSheetImport(FrappeTestCase):
    def test_import_creates_process(self):
        if frappe.db.exists("Cultivation Process", "Quy trình Sâm IMP"):
            frappe.delete_doc("Cultivation Process", "Quy trình Sâm IMP")
        rows = [
            {"Bước": 1, "Mô tả": "Chuẩn bị đất", "Công/ha": 10, "Tần suất": "1 lần/chu kỳ", "Phạm vi": "Dùng chung"},
            {"Bước": 2, "Mô tả": "Tưới nước", "Công/ha": 2, "Tần suất": "2 ngày/lần", "Phạm vi": "Theo cây"},
        ]
        name = import_rows("Quy trình Sâm IMP", "Sâm", rows)
        doc = frappe.get_doc("Cultivation Process", name)
        self.assertEqual(len(doc.steps), 2)
        self.assertEqual(doc.steps[0].scope, "shared")
        self.assertEqual(doc.steps[1].frequency_type, "n_per_period")
        self.assertEqual(doc.steps[1].frequency_value, 2)
        self.assertEqual(doc.steps[1].times_per_period, 1)

    def test_import_rows_require_photo(self):
        if frappe.db.exists("Cultivation Process", "QT Photo IMP"):
            frappe.delete_doc("Cultivation Process", "QT Photo IMP")
        rows = [
            {"Bước": 1, "Mô tả": "Có ảnh", "Tần suất": "Hàng ngày", "Phạm vi": "Dùng chung", "Yêu cầu ảnh": "x"},
            {"Bước": 2, "Mô tả": "Không ảnh", "Tần suất": "Hàng ngày", "Phạm vi": "Dùng chung", "Yêu cầu ảnh": ""},
        ]
        name = import_rows("QT Photo IMP", "Gấc", rows)
        doc = frappe.get_doc("Cultivation Process", name)
        self.assertEqual(doc.steps[0].require_photo, 1)
        self.assertEqual(doc.steps[1].require_photo, 0)


class TestParseWorkbook(FrappeTestCase):
    def test_parse_ok(self):
        name, crop, rows = sheet_import.parse_workbook(_build_wb())
        self.assertEqual(name, "Quy trình Test")
        self.assertEqual(crop, "Gấc")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["Mô tả"], "Đào hố")
        self.assertEqual(str(rows[1]["Yêu cầu ảnh"]).strip().lower(), "x")

    def test_parse_missing_name_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            sheet_import.parse_workbook(_build_wb(name=""))

    def test_parse_bad_crop_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            sheet_import.parse_workbook(_build_wb(crop="Lúa"))

    def test_parse_no_rows_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            sheet_import.parse_workbook(_build_wb(data=[]))


def _b64(wb):
    import io, base64
    bio = io.BytesIO()
    wb.save(bio)
    return base64.b64encode(bio.getvalue()).decode()


class TestImportProcessExcel(FrappeTestCase):
    def test_import_creates(self):
        if frappe.db.exists("Cultivation Process", "QT Excel A"):
            frappe.delete_doc("Cultivation Process", "QT Excel A", force=True)
        res = sheet_import.import_process_excel(_b64(_build_wb(name="QT Excel A", crop="Sâm")))
        self.assertEqual(res["exists"], False)
        self.assertEqual(res["name"], "QT Excel A")
        self.assertEqual(res["steps"], 2)
        self.assertTrue(frappe.db.exists("Cultivation Process", "QT Excel A"))

    def test_duplicate_without_replace_returns_exists(self):
        if not frappe.db.exists("Cultivation Process", "QT Excel B"):
            sheet_import.import_process_excel(_b64(_build_wb(name="QT Excel B", crop="Gấc")))
        res = sheet_import.import_process_excel(_b64(_build_wb(name="QT Excel B", crop="Gấc")))
        self.assertEqual(res["exists"], True)
        self.assertEqual(res["name"], "QT Excel B")

    def test_replace_overwrites(self):
        if not frappe.db.exists("Cultivation Process", "QT Excel C"):
            sheet_import.import_process_excel(_b64(_build_wb(name="QT Excel C", crop="Gấc")))
        one_row = [[1, "Chỉ một bước", 1, "Hàng ngày", "Dùng chung", ""]]
        res = sheet_import.import_process_excel(_b64(_build_wb(name="QT Excel C", crop="Gấc", data=one_row)), replace=1)
        self.assertEqual(res["exists"], False)
        self.assertEqual(res["steps"], 1)
        doc = frappe.get_doc("Cultivation Process", "QT Excel C")
        self.assertEqual(len(doc.steps), 1)


class TestImportOffset(FrappeTestCase):
    def test_import_rows_offset(self):
        if frappe.db.exists("Cultivation Process", "QT OFFSET IMP"):
            frappe.delete_doc("Cultivation Process", "QT OFFSET IMP", force=True)
        rows = [
            {"Bước": 1, "Mô tả": "Có offset", "Tần suất": "Hàng ngày", "Phạm vi": "Theo cây", "Bắt đầu sau (ngày)": 7},
            {"Bước": 2, "Mô tả": "Trống", "Tần suất": "Hàng ngày", "Phạm vi": "Theo cây", "Bắt đầu sau (ngày)": ""},
        ]
        name = import_rows("QT OFFSET IMP", "Gấc", rows)
        doc = frappe.get_doc("Cultivation Process", name)
        self.assertEqual(doc.steps[0].offset_days, 7)
        self.assertEqual(doc.steps[1].offset_days, -1)

    def test_template_has_offset_column(self):
        import io
        sheet_import.process_template()
        wb = openpyxl.load_workbook(io.BytesIO(frappe.response.get("filecontent")))
        ws = wb.active
        self.assertEqual(ws.cell(row=4, column=7).value, "Bắt đầu sau (ngày)")


class TestProcessTemplate(FrappeTestCase):
    def test_template_valid_and_roundtrip(self):
        import io
        sheet_import.process_template()
        content = frappe.response.get("filecontent")
        self.assertTrue(content)
        self.assertEqual(frappe.response.get("filename"), "mau-quy-trinh.xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Tên quy trình")
        self.assertEqual(ws["A2"].value, "Cây")
        self.assertEqual(
            [ws.cell(row=4, column=c).value for c in range(1, 7)],
            ["Bước", "Mô tả", "Công/ha", "Tần suất", "Phạm vi", "Yêu cầu ảnh"],
        )
        # File mẫu phải đọc được bằng chính parse_workbook
        name, crop, rows = sheet_import.parse_workbook(wb)
        self.assertEqual(crop, "Gấc")
        self.assertTrue(len(rows) >= 1)
